from openpyxl import load_workbook
wb = load_workbook('Drop_Down_Menu.xlsx')  # Work Book
ws = wb['Drop_Down_Menu']  # Work Sheet
columnA = ws['A']  # Column
columnB = ws['B']
columnC = ws["C"]
columnD = ws["D"]

column_listA = [columnA[x].value for x in range(len(columnA))]
column_listB = [columnB[x].value for x in range(len(columnB))]
column_listC = [columnC[x].value for x in range(len(columnC))]
column_listD = [columnD[x].value for x in range(len(columnD))]

print("len ", len(column_listA), " value of A ", column_listA)
print("len ", len(column_listB), " value of B ", column_listB)
print("len ", len(column_listC), " value of C ", column_listC)
print("Len ", len(column_listD), " value of D ",column_listD)


# print(column_listA[0])

list_temp_A = {}


## reading from B
if column_listA[0] != None and column_listB[0] != None:
    list_temp_A[ column_listB[0] ] = {}

for i in range(1, len(column_listA)):
    if column_listA[i] == None and column_listB[i] != None:
        list_temp_A[ column_listB[i] ] = {}

## reading from C for B#############################################################
list_temp_B_index = 0
list_temp_B = []
if column_listB[0] != None and column_listC[0] != None:
    list_temp_B.append(column_listC[0])

for i in range(1, len(column_listB)):
    if column_listB[i] is None:
        list_temp_B_index += 1
    else:
        break

temp_list  = list (column_listC)[0:list_temp_B_index]
print(" Values should be inside B", list_temp_B_index )
print("temporary column list c", temp_list, " length of B ", len(column_listB))

#####################################################################################


print(list_temp_A)


